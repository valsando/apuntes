#para meter las funciones
from openpyxl import Workbook,load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook('ejemplo.xlsx') #aqui se mete el archivo con toda la vicacion si no esta en la carpeta de python
ws = wb.active  # esta es la pagina
#print (ws['A2'].value) # imprime un valor de la tabla de la pagina actual del archivo de excel
#ws['C4'].value = 30 # para cambiar un valor de la tabla, se puede omitir el .value.
#wb.save('ejemplo.xlsx') #para guardar tiene que estar cerrado el archivo de excel.
#print (wb['Hoja1']) #pagina 1
#wb.create_sheet("test ") # crea una hoja nueva llamada test
#ws = wb.active
#ws.title = "data" # cambia el nombre de la pagina
#ws.append(["sisa","nonas", "sisa"]) #para agregar valores a las celdas en orden horizontal
#si colocas la msma apareceran mas valores en las celdas de la siuiente fila.
#wb.save("webos.xlsx")
#ws.merge_cells("A1:C1")  # para unir celdas; para despegarlas es unmerge
#    for col in range(1, 5):
#       char = chr(65 + col) # esto me dice la letra de la columna, porque a = 65, se puede hacer como ahí abajo:
#       print( ws[char + str(row)].value)  # muestra el valor de las celdas, se puede omitir el ws.value/ es importante la ubicacion del print.
#       ws[char + str(row)] = char + str(row) #esto iguala el contenido de las celdas a la letra y el numero de col y row
wb.save("ejemplo.xlsx")
